from pathlib import Path
import csv

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook


class DocumentParser:

    @staticmethod
    def read_pdf(file_path):
        pages = []
        reader = PdfReader(file_path)

        for page_number, page in enumerate(
            reader.pages,
            start=1
        ):
            page_text = page.extract_text()

            if page_text and page_text.strip():
                pages.append({
                    "page_number": page_number,
                    "text": page_text.strip()
                })

        return pages

    @staticmethod
    def read_docx(file_path):
        document = Document(file_path)

        paragraphs = []

        for paragraph in document.paragraphs:
            if paragraph.text.strip():
                paragraphs.append(
                    paragraph.text.strip()
                )

        return [{
            "page_number": None,
            "text": "\n".join(paragraphs)
        }]

    @staticmethod
    def read_txt(file_path):
        with open(
            file_path,
            "r",
            encoding="utf-8"
        ) as file:
            text = file.read()

        return [{
            "page_number": None,
            "text": text
        }]

    @staticmethod
    def normalize_cell(value):
        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def build_row_text(
        headers,
        values,
        row_number
    ):
        fields = []

        for index, value in enumerate(values):
            value = DocumentParser.normalize_cell(
                value
            )

            if not value:
                continue

            if index < len(headers):
                header = headers[index]
            else:
                header = ""

            if header:
                fields.append(
                    f"{header}: {value}"
                )
            else:
                fields.append(value)

        if not fields:
            return None

        return (
            f"Row {row_number}: "
            + " | ".join(fields)
        )

    @staticmethod
    def read_xlsx(file_path):
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True
        )

        sections = []

        try:
            for worksheet in workbook.worksheets:
                rows = worksheet.iter_rows(
                    values_only=True
                )

                try:
                    first_row = next(rows)
                except StopIteration:
                    continue

                headers = [
                    DocumentParser.normalize_cell(
                        value
                    )
                    for value in first_row
                ]

                # If the first row is completely empty,
                # treat the sheet as having no headers.
                has_headers = any(headers)

                row_texts = []

                if not has_headers:
                    first_text = (
                        DocumentParser.build_row_text(
                            [],
                            first_row,
                            1
                        )
                    )

                    if first_text:
                        row_texts.append(first_text)

                for row_number, row in enumerate(
                    rows,
                    start=2
                ):
                    row_text = (
                        DocumentParser.build_row_text(
                            headers
                            if has_headers
                            else [],
                            row,
                            row_number
                        )
                    )

                    if row_text:
                        row_texts.append(
                            row_text
                        )

                if not row_texts:
                    continue

                text = (
                    f"Sheet: {worksheet.title}\n"
                    + "\n".join(row_texts)
                )

                sections.append({
                    "page_number": None,
                    "sheet_name": worksheet.title,
                    "text": text
                })

        finally:
            workbook.close()

        return sections

    @staticmethod
    def read_csv(file_path):
        rows_output = []

        with open(
            file_path,
            "r",
            encoding="utf-8-sig",
            newline=""
        ) as file:
            reader = csv.reader(file)

            try:
                first_row = next(reader)
            except StopIteration:
                return []

            headers = [
                DocumentParser.normalize_cell(
                    value
                )
                for value in first_row
            ]

            has_headers = any(headers)

            if not has_headers:
                first_text = (
                    DocumentParser.build_row_text(
                        [],
                        first_row,
                        1
                    )
                )

                if first_text:
                    rows_output.append(
                        first_text
                    )

            for row_number, row in enumerate(
                reader,
                start=2
            ):
                row_text = (
                    DocumentParser.build_row_text(
                        headers
                        if has_headers
                        else [],
                        row,
                        row_number
                    )
                )

                if row_text:
                    rows_output.append(
                        row_text
                    )

        if not rows_output:
            return []

        return [{
            "page_number": None,
            "sheet_name": None,
            "text": "\n".join(rows_output)
        }]

    @staticmethod
    def parse(file_path):
        extension = (
            Path(file_path)
            .suffix
            .lower()
        )

        if extension == ".pdf":
            return DocumentParser.read_pdf(
                file_path
            )

        elif extension == ".docx":
            return DocumentParser.read_docx(
                file_path
            )

        elif extension == ".txt":
            return DocumentParser.read_txt(
                file_path
            )

        elif extension == ".xlsx":
            return DocumentParser.read_xlsx(
                file_path
            )

        elif extension == ".csv":
            return DocumentParser.read_csv(
                file_path
            )

        else:
            raise ValueError(
                f"Unsupported file type: "
                f"{extension}"
            )